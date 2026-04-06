from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .utils import ensure_dir


def _load_existing_share_data(report_path: Path) -> dict[str, tuple[str, str]]:
    if not report_path.exists():
        return {}
    workbook = load_workbook(report_path)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    try:
        docx_index = header.index("docx_file") + 1
        share_link_index = header.index("share_link") + 1
        share_notes_index = header.index("share_notes") + 1
    except ValueError:
        return {}

    share_data: dict[str, tuple[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=False):
        docx_path = str(row[docx_index - 1].value or "")
        if not docx_path:
            continue
        share_link = str(row[share_link_index - 1].value or "")
        share_notes = str(row[share_notes_index - 1].value or "")
        share_data[docx_path] = (share_link, share_notes)
    return share_data


def _set_hyperlink(cell, target: str, label: str | None = None) -> None:
    if not target:
        cell.value = ""
        return
    cell.value = label or target
    cell.hyperlink = target
    cell.font = Font(color="0563C1", underline="single")


def write_resume_tracker(rows: list[dict[str, str]], output_dir: Path) -> Path:
    ensure_dir(output_dir)
    report_path = output_dir / "resume_tracker.xlsx"
    existing_share_data = _load_existing_share_data(report_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumes"
    headers = [
        "processed_at",
        "company",
        "role",
        "job_link",
        "markdown_file",
        "docx_file",
        "share_link",
        "share_notes",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row_index, item in enumerate(rows, start=2):
        docx_path = item["docx_path"]
        share_link, share_notes = existing_share_data.get(
            docx_path,
            ("", "Upload the DOCX or PDF to Drive/OneDrive and paste the public link here."),
        )
        sheet.cell(row=row_index, column=1, value=item["processed_at"])
        sheet.cell(row=row_index, column=2, value=item["company"])
        sheet.cell(row=row_index, column=3, value=item["title"])
        _set_hyperlink(sheet.cell(row=row_index, column=4), item["apply_url"], "Open job")

        markdown_target = Path(item["markdown_path"]).resolve() if item["markdown_path"] else None
        docx_target = Path(docx_path).resolve() if docx_path else None
        _set_hyperlink(
            sheet.cell(row=row_index, column=5),
            markdown_target.as_uri() if markdown_target else "",
            str(markdown_target) if markdown_target else "",
        )
        _set_hyperlink(
            sheet.cell(row=row_index, column=6),
            docx_target.as_uri() if docx_target else "",
            str(docx_target) if docx_target else "",
        )
        sheet.cell(row=row_index, column=7, value=share_link)
        sheet.cell(row=row_index, column=8, value=share_notes)

    widths = {
        1: 24,
        2: 18,
        3: 48,
        4: 14,
        5: 60,
        6: 60,
        7: 45,
        8: 60,
    }
    for column_index, width in widths.items():
        sheet.column_dimensions[chr(64 + column_index)].width = width

    sheet.freeze_panes = "A2"
    workbook.save(report_path)
    return report_path
